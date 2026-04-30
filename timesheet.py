"""Timesheet export of user-typed messages, formatted for hand-reconstructing a timesheet."""

import csv
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NY = ZoneInfo("America/New_York")
UTC = ZoneInfo("UTC")

HEADERS = [
    "date", "weekday", "time_24h", "project", "session_title", "session_id",
    "minutes_since_prev_message", "gap_since_prev_message",
    "message_char_count", "message",
]

# Pastel palette — readable with default black text.
PALETTE = [
    "FFE4B5", "B5E7D8", "C8D8F0", "F4C2C2", "E0CFF0",
    "FFF1A8", "C8E6C9", "FFD6A5", "BFE3F0", "E6D7C3",
    "F8C8DC", "D5E8B5", "C7CEEA", "F2D7B5", "B5D7F2",
]

COLUMN_WIDTHS = {
    "date": 12, "weekday": 11, "time_24h": 10, "project": 24,
    "session_title": 28, "session_id": 38,
    "minutes_since_prev_message": 12, "gap_since_prev_message": 14,
    "message_char_count": 8, "message": 90,
}


def _human_gap(total_minutes: float) -> str:
    total = int(round(total_minutes))
    if total < 1:
        return "<1m"
    days, rem = divmod(total, 60 * 24)
    hours, minutes = divmod(rem, 60)
    parts = []
    if days:
        parts.append(f"{days}d")
    if hours:
        parts.append(f"{hours}h")
    if minutes or not parts:
        parts.append(f"{minutes}m")
    return " ".join(parts)


def fetch_rows(db_path: Path, days: int = 4) -> list[dict]:
    """Return user-typed message rows for the past `days` NY-local days (today + days-1 prior).

    Each day starts at midnight NY-local. The `minutes_since_prev_message` gap is computed
    against the previous message in any project, using a 2-day pre-window lookback so the
    very first row in the window still has an accurate gap.
    """
    today_ny = datetime.now(NY).date()
    start_dt_ny = datetime.combine(today_ny - timedelta(days=days - 1),
                                   datetime.min.time(), tzinfo=NY)
    end_dt_ny = datetime.combine(today_ny + timedelta(days=1),
                                 datetime.min.time(), tzinfo=NY)
    lookback_utc = (start_dt_ny - timedelta(days=2)).astimezone(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    end_utc = end_dt_ny.astimezone(UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    raw = conn.execute("""
        SELECT m.timestamp, m.project, m.session_id, m.content,
               COALESCE(NULLIF(s.name, ''), NULLIF(s.slug, ''), '') AS session_title
        FROM messages m
        LEFT JOIN sessions s ON s.session_id = m.session_id
        WHERE m.timestamp >= ? AND m.timestamp < ?
          AND m.content NOT LIKE '<task-notification>%'
          AND m.content NOT LIKE '<local-command-stdout>%'
          AND m.content NOT LIKE '<bash-stdout>%'
        ORDER BY m.timestamp ASC
    """, (lookback_utc, end_utc)).fetchall()
    conn.close()

    prev_dt = None
    out = []
    for r in raw:
        ts_utc = datetime.strptime(r["timestamp"].replace("Z", "+0000"),
                                   "%Y-%m-%dT%H:%M:%S.%f%z")
        ts_ny = ts_utc.astimezone(NY)
        minutes_since_prev = ""
        gap_human = ""
        if prev_dt is not None:
            delta_min = (ts_utc - prev_dt).total_seconds() / 60.0
            minutes_since_prev = round(delta_min, 1)
            gap_human = _human_gap(delta_min)
        prev_dt = ts_utc
        if ts_ny < start_dt_ny or ts_ny >= end_dt_ny:
            continue
        out.append({
            "date": ts_ny.strftime("%Y-%m-%d"),
            "weekday": ts_ny.strftime("%A"),
            "time_24h": ts_ny.strftime("%H:%M:%S"),
            "project": r["project"] or "",
            "session_title": r["session_title"],
            "session_id": r["session_id"],
            "minutes_since_prev_message": minutes_since_prev,
            "gap_since_prev_message": gap_human,
            "message_char_count": len(r["content"] or ""),
            "message": r["content"] or "",
        })
    return out


def write_csv(rows: list[dict], path: Path) -> None:
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict], path: Path) -> None:
    projects = []
    for r in rows:
        if r["project"] not in projects:
            projects.append(r["project"])
    project_fill = {
        p: PatternFill(start_color=PALETTE[i % len(PALETTE)],
                       end_color=PALETTE[i % len(PALETTE)],
                       fill_type="solid")
        for i, p in enumerate(projects)
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="top")

    for r in rows:
        ws.append([r[h] for h in HEADERS])
        row_idx = ws.max_row
        fill = project_fill.get(r["project"])
        for col_idx, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(header == "message"))

    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS.get(h, 14)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    legend = wb.create_sheet("Project Colors")
    legend.append(["project", "color"])
    for cell in legend[1]:
        cell.font = Font(bold=True)
    for p in projects:
        legend.append([p, ""])
        legend.cell(row=legend.max_row, column=2).fill = project_fill[p]
    legend.column_dimensions["A"].width = 30
    legend.column_dimensions["B"].width = 12

    wb.save(path)
